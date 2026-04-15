from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FILE = ROOT / "data" / "Nguyenvong.xlsx"

wb = load_workbook(FILE, data_only=True)
for sheet_name in [name for name in ("Sheet1", "Sheet2") if name in wb.sheetnames]:
    ws = wb[sheet_name]
    print(f"=== {sheet_name} ===")
    for r in range(1, 8):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        print(r, vals)
