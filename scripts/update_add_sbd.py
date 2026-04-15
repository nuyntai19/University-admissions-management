from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

DS_FILE = ROOT / "data" / "Ds thi sinh.xlsx"
NV_FILE = ROOT / "data" / "Nguyenvong.xlsx"
BACKUP = ROOT / "data" / "Nguyenvong.backup.xlsx"


# =========================
# BUILD MAP CCCD -> SBD
# =========================
def build_map():
    wb = load_workbook(DS_FILE, data_only=True)
    ws = wb.active

    # dò header dòng 1-10
    header_row = None
    for r in range(1, 20):
        row = [str(c.value).strip().lower() if c.value else "" for c in ws[r]]
        if "cccd" in row and ("số báo danh" in row or "sbd" in row):
            header_row = r
            break

    if not header_row:
        header_row = 1

    headers = [c.value for c in ws[header_row]]

    cccd_col = None
    sbd_col = None

    for i, h in enumerate(headers):
        if not h:
            continue
        h = str(h).lower()
        if "cccd" in h:
            cccd_col = i
        if "số báo danh" in h or "sbd" in h:
            sbd_col = i

    if cccd_col is None or sbd_col is None:
        raise Exception("Không tìm thấy CCCD / SBD trong Ds thi sinh.xlsx")

    mapping = {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue

        cccd = str(row[cccd_col]).strip() if row[cccd_col] else ""
        sbd = str(row[sbd_col]).strip() if row[sbd_col] else ""

        if cccd and sbd:
            mapping[cccd] = sbd

    return mapping


# =========================
# UPDATE NGUYENVONG
# =========================
def update_file():
    print("Starting update...")

    mapping = build_map()
    print(f"Loaded {len(mapping)} mappings")

    wb = load_workbook(NV_FILE)
    ws = wb.active

    wb.save(BACKUP)
    print(f"Backup saved: {BACKUP}")

    # ===== FIX CỐ ĐỊNH HEADER ROW = 5 (theo file bạn gửi) =====
    header_row = 5
    headers = [c.value for c in ws[header_row]]

    cccd_col = None

    for i, h in enumerate(headers):
        if h and "cccd" in str(h).lower():
            cccd_col = i + 1
            break

    if cccd_col is None:
        raise Exception("Không tìm thấy cột CCCD trong Nguyenvong.xlsx")

    # insert SBD column
    ws.insert_cols(cccd_col + 1)
    ws.cell(row=header_row, column=cccd_col + 1).value = "Số báo danh"

    updated = 0

    for r in range(header_row + 1, ws.max_row + 1):
        cccd = ws.cell(r, cccd_col).value
        if not cccd:
            continue

        cccd = str(cccd).strip()
        sbd = mapping.get(cccd)

        if sbd:
            ws.cell(r, cccd_col + 1).value = sbd
            updated += 1

    wb.save(NV_FILE)

    print(f"Done! Updated {updated} rows")


if __name__ == "__main__":
    update_file()