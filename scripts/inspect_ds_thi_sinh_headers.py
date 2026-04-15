from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FILE = ROOT / "data" / "Ds thi sinh.xlsx"

wb = load_workbook(FILE, data_only=True)
ws = wb.active
print(f"Sheet: {ws.title}")
for r in range(1, 10):
    vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 25) + 1)]
    print(r, vals)
