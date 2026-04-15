from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FILE = ROOT / "data" / "Nguyenvong.xlsx"
TS_PATTERN = re.compile(r"^TS_\d+$", re.IGNORECASE)


def remove_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def detect_header_row(ws, max_scan_rows: int = 20) -> int:
    best_row = 1
    best_score = -1
    keywords = ["cccd", "thu tu", "ma xet tuyen", "ma truong"]

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        row_text = " | ".join(remove_accents(str(v)).lower() for v in row_vals if v is not None)
        score = sum(1 for kw in keywords if kw in row_text)
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row

wb = load_workbook(FILE, data_only=True)
for sheet_name in [name for name in ("Sheet1", "Sheet2") if name in wb.sheetnames]:
    ws = wb[sheet_name]
    header_row = detect_header_row(ws)
    headers = [
        str(c).strip() if c is not None else ""
        for c in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    ]
    cccd_col = None
    for i, h in enumerate(headers, start=1):
        hl = remove_accents(h).lower()
        if "cccd" in hl or "can cuoc" in hl or "cmnd" in hl:
            cccd_col = i
            break

    if cccd_col is None:
        print(f"[{sheet_name}] CCCD column not found")
        continue

    ts_count = 0
    digit12_count = 0
    non_empty = 0
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, cccd_col).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        non_empty += 1
        if TS_PATTERN.match(s):
            ts_count += 1
        if s.isdigit() and len(s) == 12:
            digit12_count += 1

    print(f"[{sheet_name}] non-empty={non_empty}, still_TS={ts_count}, cccd_12_digits={digit12_count}")
