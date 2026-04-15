from __future__ import annotations

import re
import shutil
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DS_THI_SINH_FILE = DATA_DIR / "Ds thi sinh.xlsx"
NGUYEN_VONG_FILE = DATA_DIR / "Nguyenvong.xlsx"
BACKUP_FILE = DATA_DIR / "Nguyenvong.backup.xlsx"

TS_PATTERN = re.compile(r"^TS_\d+$", re.IGNORECASE)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def remove_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def normalize_sbd(value: object) -> str:
    raw = normalize_text(value).upper()
    if not raw:
        return ""
    return re.sub(r"\s+", "", raw)


def normalize_cccd(value: object) -> str:
    if value is None:
        return ""

    if isinstance(value, (int, float)):
        int_value = int(value)
        return str(int_value).zfill(12)

    raw = normalize_text(value)
    if not raw:
        return ""

    raw_no_space = re.sub(r"\s+", "", raw)
    if raw_no_space.isdigit():
        return raw_no_space.zfill(12)

    return raw_no_space


def score_header(header_name: str, keywords: Iterable[str]) -> int:
    normalized = remove_accents(header_name).lower().replace("_", " ").replace("-", " ")
    score = 0
    for kw in keywords:
        if kw in normalized:
            score += 1
    return score


def find_header_row(ws, max_scan_rows: int = 20) -> Optional[int]:
    best_row = None
    best_score = 0

    expected_keywords = ["cccd", "so bao danh", "sbd", "thu tu", "ma xet tuyen"]

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_values = [normalize_text(c.value) for c in ws[row_idx]]
        row_text = " | ".join(remove_accents(v).lower() for v in row_values if v)
        if not row_text:
            continue

        score = sum(1 for kw in expected_keywords if kw in row_text)
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def find_column_indexes(headers: Dict[int, str]) -> Tuple[Optional[int], Optional[int]]:
    # Heuristic to locate SBD and CCCD columns from header row.
    sbd_keywords = ["sbd", "so bao danh", "bao danh", "mats", "ma thi sinh", "thi sinh"]
    cccd_keywords = ["cccd", "can cuoc", "cmnd", "so cccd", "so cmnd"]

    best_sbd_col = None
    best_sbd_score = 0
    best_cccd_col = None
    best_cccd_score = 0

    for col_idx, header in headers.items():
        name = normalize_text(header)
        if not name:
            continue

        sbd_score = score_header(name, sbd_keywords)
        cccd_score = score_header(name, cccd_keywords)

        if sbd_score > best_sbd_score:
            best_sbd_score = sbd_score
            best_sbd_col = col_idx

        if cccd_score > best_cccd_score:
            best_cccd_score = cccd_score
            best_cccd_col = col_idx

    return best_sbd_col, best_cccd_col


def build_sbd_to_cccd_map() -> Dict[str, str]:
    if not DS_THI_SINH_FILE.exists():
        raise FileNotFoundError(f"Missing file: {DS_THI_SINH_FILE}")

    wb = load_workbook(DS_THI_SINH_FILE, data_only=True)
    ws = wb.active

    header_row_idx = find_header_row(ws) or 1
    header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    headers = {idx + 1: value for idx, value in enumerate(header_row)}
    sbd_col, cccd_col = find_column_indexes(headers)

    if not sbd_col or not cccd_col:
        raise ValueError(
            "Could not detect SBD/CCCD columns in Ds thi sinh.xlsx. "
            f"Detected columns: {headers}"
        )

    result: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        sbd = normalize_sbd(row[sbd_col - 1])
        cccd = normalize_cccd(row[cccd_col - 1])
        if not sbd or not cccd:
            continue
        if not TS_PATTERN.match(sbd):
            continue
        if sbd not in result:
            result[sbd] = cccd

    return result


def update_nguyenvong(sbd_to_cccd: Dict[str, str]) -> None:
    if not NGUYEN_VONG_FILE.exists():
        raise FileNotFoundError(f"Missing file: {NGUYEN_VONG_FILE}")

    if not BACKUP_FILE.exists():
        shutil.copy2(NGUYEN_VONG_FILE, BACKUP_FILE)

    wb = load_workbook(NGUYEN_VONG_FILE)

    target_sheet_names = [name for name in ("Sheet1", "Sheet2") if name in wb.sheetnames]
    if not target_sheet_names:
        target_sheet_names = wb.sheetnames[:2]

    total_updated = 0
    total_missing = 0

    for sheet_name in target_sheet_names:
        ws = wb[sheet_name]
        header_row_idx = find_header_row(ws) or 1
        header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
        headers = {idx + 1: value for idx, value in enumerate(header_row)}

        _, cccd_col = find_column_indexes(headers)
        if not cccd_col:
            print(f"[{sheet_name}] Could not detect CCCD column. Skipped.")
            continue

        updated_count = 0
        missing_count = 0

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=cccd_col)
            raw_value = normalize_text(cell.value)
            lookup_key = normalize_sbd(raw_value)

            if not TS_PATTERN.match(lookup_key):
                continue

            mapped_cccd = sbd_to_cccd.get(lookup_key)
            if mapped_cccd:
                cell.value = mapped_cccd
                updated_count += 1
            else:
                missing_count += 1

        print(f"[{sheet_name}] Updated: {updated_count}, Missing SBD in Ds thi sinh: {missing_count}")
        total_updated += updated_count
        total_missing += missing_count

    wb.save(NGUYEN_VONG_FILE)
    print(f"Done. Total updated: {total_updated}, Total missing: {total_missing}")
    print(f"Backup file: {BACKUP_FILE}")


def main() -> None:
    sbd_to_cccd = build_sbd_to_cccd_map()
    print(f"Loaded {len(sbd_to_cccd)} SBD->CCCD mappings from Ds thi sinh.xlsx")
    update_nguyenvong(sbd_to_cccd)


if __name__ == "__main__":
    main()
